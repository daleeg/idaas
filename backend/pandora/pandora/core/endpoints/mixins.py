# -*- coding: utf-8 -*-
from __future__ import absolute_import, unicode_literals
import logging
from django.db import transaction
from rest_framework import mixins
from rest_framework import status
from rest_framework.decorators import action
from django.http.response import HttpResponse
from django.utils.translation import ugettext_lazy as _
from django.utils.encoding import escape_uri_path
from openpyxl import Workbook, load_workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Border, Font, Side, Alignment, PatternFill
from pandora.utils.decorators import swagger_auto_schema
from pandora.core.response import APIResponse
from pandora.core.code import BAD_REQUEST
from pandora.core.exceptions import DeleteProtectedError
from pandora.core.schema import default_parameters, list_parameters, retrieve_parameters
from pandora.core.schema import DestroySerializer, BulkDestroySerializer

LOG = logging.getLogger(__name__)


class CreateModelMixin(mixins.CreateModelMixin):
    """
    Create a model instance.
    """

    def perform_create(self, serializer, **kwargs):
        instance = serializer.save(**kwargs)
        return serializer

    @swagger_auto_schema(manual_parameters=default_parameters)
    def create(self, request, *args, **kwargs):
        data = request.data
        try:
            with transaction.atomic():
                result = self._create(data)

        except Exception as e:
            LOG.error(e)
            raise
        headers = self.get_success_headers(data)
        return APIResponse(data=result, status=status.HTTP_201_CREATED, headers=headers)

    def _create(self, data):
        data = self.pre_request_data(data)
        serializer = self.get_serializer(data=data)
        serializer.is_valid(raise_exception=True)
        serializer = self.perform_create(serializer)
        return serializer.data


class BulkCreateModelMixin(mixins.CreateModelMixin):
    def perform_bulk_create(self, serializer, **kwargs):
        instances = serializer.save(**kwargs)
        return serializer

    @swagger_auto_schema(manual_parameters=default_parameters)
    def bulk_create(self, request, *args, **kwargs):
        data = request.data
        try:
            result = self._bulk_create(data)
        except Exception as e:
            LOG.error(e)
            raise
        headers = self.get_success_headers(data)
        return APIResponse(data=result, status=status.HTTP_201_CREATED, headers=headers)

    def _bulk_create(self, data):
        result = []
        for item in data:
            with transaction.atomic():
                item = self.pre_request_data(item)
                serializer = self.get_serializer(data=item)
                serializer.is_valid(raise_exception=True)
                serializer = self.perform_bulk_create(serializer)
                result.append(serializer.data)
        return result


class RetrieveModelMixin(mixins.RetrieveModelMixin):
    """
    Retrieve a model instance.
    """
    lookup_field = "uid"

    @swagger_auto_schema(manual_parameters=retrieve_parameters)
    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        serializer = self.get_serializer(instance)
        return APIResponse(data=self.final_response_data(serializer.data))


class UpdateModelMixin(mixins.UpdateModelMixin):
    """
     Update a model instance.
     """

    lookup_field = "uid"

    @swagger_auto_schema(manual_parameters=default_parameters)
    def update(self, request, *args, **kwargs):
        partial = kwargs.pop("partial", True)
        instance = self.get_object()
        try:
            with transaction.atomic():
                serializer = self.get_serializer(instance, data=self.pre_update_request_data(instance, request.data),
                                                 partial=partial)
                serializer.is_valid(raise_exception=True)
                serializer = self.perform_update(serializer)
        except Exception as e:
            LOG.error(e)
            raise

        if getattr(instance, "_prefetched_objects_cache", None):
            # If "prefetch_related" has been applied to a queryset, we need to
            # forcibly invalidate the prefetch cache on the instance.
            instance._prefetched_objects_cache = {}

        return APIResponse(data=self.final_response_data(serializer.data))

    def perform_update(self, serializer):
        instance = serializer.save()
        return serializer


class DestroyModelMixin(mixins.DestroyModelMixin):
    """
    Destroy a model instance.
    """
    lookup_field = "uid"
    protected_related_fields = ()
    default_soft = True

    @swagger_auto_schema(manual_parameters=default_parameters, request_body=DestroySerializer)
    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        with transaction.atomic():
            self.perform_destroy(instance)

        return APIResponse()

    @property
    def soft_delete(self):
        soft = self.request.data.get("soft", self.default_soft)
        return soft

    def perform_destroy(self, instance):
        self.pre_destroy(instance)
        soft = self.soft_delete
        instance.delete(soft=soft)

    def pre_destroy(self, instance):
        if not self.request.data.get("force", False):
            for field in self.protected_related_fields:
                if getattr(instance, field).all().exists():
                    model_name = instance.__class__.__name__
                    # model_id = getattr(ModelSet, model_name)
                    # if model_id:
                    #     model_name = ModelSet.MESSAGE.get(model_id, model_name)
                    raise DeleteProtectedError(model_name, field)


class BulkDestroyModelMixin(object):
    @swagger_auto_schema(manual_parameters=default_parameters, request_body=BulkDestroySerializer)
    def bulk_destroy(self, request, *args, **kwargs):
        # queryset = self.filter_queryset(self.get_queryset())
        queryset = self.get_queryset()
        param = self.request.query_params.copy()
        if param:
            param = param.dict()
            queryset = queryset.filter(**param).distinct()
        ids = []
        if request.data:
            data = request.data.copy()
            ids_str = data.get("ids", "")
            if isinstance(ids_str, str):
                ids = ids_str.split(",")
            else:
                ids = ids_str
            if ids:
                instances = queryset.filter(uid__in=ids)
            else:
                instances = queryset
        else:
            instances = queryset
        try:
            if not self.request.query_params and not ids:
                return APIResponse(code=BAD_REQUEST, message="不能删除所有数据")
            if not instances.exists():
                return APIResponse()

            with transaction.atomic():
                self.perform_bat_destroy(instances)

            return APIResponse()
        except DeleteProtectedError:
            raise
        except Exception as e:
            return APIResponse(code=BAD_REQUEST, message="delete failed:, details{}".format(e))

    def perform_bat_destroy(self, instances):
        for instance in instances:
            self.perform_destroy(instance)


class ListModelMixin(object):
    """
    List a queryset.
    """

    @swagger_auto_schema(manual_parameters=list_parameters)
    def list(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.get_distinct():
            queryset = queryset.distinct()
        if self.order_by_fields and isinstance(self.order_by_fields, tuple):
            queryset = queryset.order_by(*self.order_by_fields)
        page = self.paginate_queryset(queryset)
        if page is not None:
            serializer = self.get_list_serializer(page, many=True)
            return self.get_paginated_response(self.final_response_data(serializer.data))

        serializer = self.get_list_serializer(queryset, many=True)

        return APIResponse(data=self.final_response_data(serializer.data))

    def get_list_serializer(self, *args, **kwargs):
        return self.get_serializer(*args, **kwargs)


class DownloadExcelModelMixin(object):
    explain_list = []
    data_list = []
    split_sign = ","
    multi = False
    sheet_list = None
    add_extra = False
    output_name = ""
    ext = "xlsx"

    def download(self, request, *args, **kwargs):
        queryset = self.filter_queryset(self.get_queryset())
        if self.get_distinct():
            queryset = queryset.distinct()
        if self.order_by_fields and isinstance(self.order_by_fields, tuple):
            queryset = queryset.order_by(*self.order_by_fields)
        self.set_data(queryset, *args, **kwargs)
        response = self.get_response(request)
        return response

    @NotImplementedError
    def set_data(self, queryset):
        """
        set excel data
        # multi: 是否有多个sheet
        # sheet_list: 多sheet的名称
        # data_list: 表数据,第一个列表元素为表头[[headers], [data1], [data2]]。
                     如果multi的话,[[[headers], [data1], [data2]], [[headers], [data1], [data2]]]
        # output_name: 输出文件名
        # add_extra: 是否增加额外信息
        # explain_list: 声明提示信息
        """
        pass

    @property
    def content(self):
        workbook = Workbook()
        for index, sheet_name in enumerate(self.sheet_list or ["Sheet1"]):
            data = self.data_list[index] if self.multi else self.data_list
            # if not self.multi:
            #     data = [x.split(self.split_sign) for x in self.data_list]
            # else:
            #     data = [x.split(self.split_sign) for x in self.data_list[index]]

            if index > 0:
                sheet = workbook.create_sheet(title=sheet_name)
                workbook.active = sheet
            else:
                sheet = workbook.active
                sheet.title = sheet_name

            border = Border(*([Side(style="thin", color="000000")] * 4))

            rows, columns = len(data), len(data[0])
            if self.explain_list:
                self.write_explain(sheet, columns, 0)
            explain_rows = len(self.explain_list)
            max_width, row_index = 0, 0
            for row in range(explain_rows, rows + explain_rows):
                row_index += 1
                sheet.row_dimensions[row + 1].height = 18
                for column in range(columns):
                    the_cell = sheet["%s%s" % (chr(column + 65), row + 1)]
                    value = data[row][column].strip()
                    the_cell.value, the_cell.border = value, border
                    if row_index == 1:
                        the_cell.font = Font(b=True, color="FFFFFF")
                        the_cell.fill = PatternFill(fill_type="solid", fgColor="6A5ACD")
                    if row_index > 1:
                        if column == 3:
                            if "\n" in the_cell.value:
                                the_cell.alignment = Alignment(horizontal="left", vertical="top")
                            else:
                                the_cell.alignment = Alignment(horizontal="left", vertical="center")
                                if len(the_cell.value) > max_width:
                                    max_width = len(the_cell.value)
                        elif column == columns - 1:
                            the_cell.alignment = Alignment(horizontal="center", vertical="center")
                            the_cell.number_format = "0.00  "
                        else:
                            the_cell.alignment = Alignment(horizontal="center", vertical="center")
                    else:
                        the_cell.alignment = Alignment(horizontal="center", vertical="center")
            for column in range(columns):
                # sheet["%s1" % chr(column + 65)].font = Font(b=True, color="FFFFFF")
                column_width = 20 if column < 3 else (12 if column > 3 else (int(max_width * 1.58) or 30))
                sheet.column_dimensions[chr(column + 65)].width = column_width

        workbook.active = 0
        workbook.close()
        return save_virtual_workbook(workbook)

    def write_explain(self, ws, column_num, row):
        for index, explain in enumerate(self.explain_list):
            current_row = row + index
            for column_index in range(column_num):
                current_cell = ws["{}{}".format(chr(column_index + 65), current_row)]
                current_cell.value = explain if column_index == 1 else ""
                current_cell.fill = PatternFill(fill_type="solid", fgColor="FFFF00")

    def get_response(self, request):
        content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        response = HttpResponse(content=self.content, content_type=content_type)
        response["Content-Disposition"] = "attachment; filename*=utf-8""{}".format(
            escape_uri_path("{}.{}".format(self.output_name, self.ext)))
        return response


class UploadExcelModelMixin(object):
    model_name = None
    sheet_map = {}
    explain_rows = 0  # 说明行数

    def define_row_type(self, row, type_list):
        if len(row) != len(type_list):
            raise ValueError("row和type_list长度不同")
        for index in range(len(row)):
            row[index] = self.define_type(row[index], type_list[index])

    @staticmethod
    def define_type(value, type_class):
        return type_class(value) if value else value

    def upload(self, request, *args, **kwargs):
        """导入"""
        trunk_file = request.FILES.get("file", None)
        if not trunk_file:
            return APIResponse(code=BAD_REQUEST, message=_("未读取到文件"))
        try:
            wb = load_workbook(filename=trunk_file)
            work_sheets = wb.worksheets
        except Exception as e:
            LOG.error(e)
            return APIResponse(code=BAD_REQUEST, message=_("Excel 出错"))

        self.load_params(request)
        for sheet in work_sheets:
            title = sheet.title
            if title not in self.sheet_map:
                if len(self.sheet_map) > 1:
                    LOG.info("not support sheet: {}".format(title))
                    continue

            data, messages = self.load_data(title, sheet)
            if messages:
                return APIResponse(code=BAD_REQUEST, message=messages)

            if not data:
                return APIResponse()

            self.save_data(title, data)
            LOG.info("upload sheet : {} finish".format(title))
        return APIResponse()

    def load_params(self, request):
        pass

    def load_data(self, title, sheet):
        is_header = True
        result = []
        row_index = 0
        for row in sheet.rows:
            row_index += 1
            if row_index <= self.explain_rows:
                continue
            if is_header:
                is_header = False
                continue
            item = [i.value for i in row]
            result.append(item)
        return self.reduce_data(title, result)

    def reduce_data(self, title, rows):
        table = self.sheet_map.get(title, self.model_name)
        func = getattr(self, "reduce_{}".format(table), None)
        if not func:
            LOG.warning("not register reduce func: {}".format(title))
            return
        return func(title, rows)

    def save_data(self, title, result):
        table = self.sheet_map.get(title, self.model_name)
        func = getattr(self, "save_{}".format(table))
        if not func:
            LOG.warning("not register save func: {}".format(title))
            return
        return func(title, result)
